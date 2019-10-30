import RPi.GPIO as GPIO
from py532lib.i2c import *
from py532lib.frame import *
from py532lib.constants import *
from openpyxl import workbook
from openpyxl import load_workbook
from datetime import datetime


#Buzzer
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BOARD)
buzzer=40
GPIO.setup(buzzer,GPIO.OUT)

#NFC
pn532 = Pn532_i2c()
pn532.SAMconfigure()

#Shared File
#mount -t cifs //192.168.1.2/share /mnt/share -o username=admin
file='/mnt/share/clock.xlsx'
wb = load_workbook(file)
sheets = wb.sheetnames
sheet0 = wb[sheets[0]]
r=2

#Employees
id,nom=('188','128'),('toto','titi')


#Register
while True:
        GPIO.output(buzzer,GPIO.HIGH)
        sleep(1)
        GPIO.output(buzzer,GPIO.LOW)

        card_data = pn532.read_mifare().get_data()

        time = datetime.now()
        date = time.strftime("%d/%m/%Y")
        hour = time.strftime("%H:%M:%S")

        for i in range(len(id)):
                #Tag id
		if str(card_data[7]) == id[i]:
			print(nom[i]," arrives at : ", date,hour)
                	if sheet0.cell(row = r, column = 1).value is None:
                        	sheet0.cell(row = r, column = 1).value = nom[i]
	                        sheet0.cell(row = r, column = 2).value = date
	                        sheet0.cell(row = r, column = 3).value = hour
        	                wb.save(file)
			else:
                        	r+=1
	                        sheet0.cell(row = r, column = 1).value = nom[i]
	                        sheet0.cell(row = r, column = 2).value = date
	                        sheet0.cell(row = r, column = 3).value = hour
	                        wb.save(file)
